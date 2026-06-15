import os
import re
import json
from datetime import datetime, date, timedelta
from flask import Flask, request, jsonify, render_template, g
import openpyxl
import psycopg2
import psycopg2.extras
from dotenv import load_dotenv

load_dotenv()

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 20 * 1024 * 1024

DATABASE_URL = os.environ.get('DATABASE_URL')

TYPES = [
    'Customer Attributed (Compute)',
    'Customer Specific (Storage,Read/write)',
    'Platform',
]

# ---------------------------------------------------------------------------
# DB helpers
# ---------------------------------------------------------------------------

def get_db():
    if 'db' not in g:
        conn = psycopg2.connect(DATABASE_URL)
        conn.autocommit = False
        g.db = conn
    return g.db

def get_cur(db):
    return db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

@app.teardown_appcontext
def close_db(e=None):
    db = g.pop('db', None)
    if db and not db.closed:
        db.close()

def init_db():
    conn = psycopg2.connect(DATABASE_URL)
    cur = conn.cursor()
    cur.execute('''
        CREATE TABLE IF NOT EXISTS resource_type_map (
            resource_key TEXT PRIMARY KEY,
            type         TEXT NOT NULL,
            source       TEXT DEFAULT 'user'
        )
    ''')
    cur.execute('''
        CREATE TABLE IF NOT EXISTS daily_costs (
            id                SERIAL PRIMARY KEY,
            upload_date       TEXT NOT NULL,
            resource          TEXT NOT NULL,
            resource_id       TEXT,
            resource_type     TEXT,
            resource_group    TEXT,
            subscription_name TEXT,
            cost_inr          REAL NOT NULL,
            cost_usd          REAL,
            currency          TEXT,
            type              TEXT NOT NULL
        )
    ''')
    cur.execute('''
        CREATE TABLE IF NOT EXISTS pending_classifications (
            session_id        TEXT NOT NULL,
            resource          TEXT NOT NULL,
            resource_id       TEXT,
            resource_type     TEXT,
            resource_group    TEXT,
            subscription_name TEXT,
            cost_inr          REAL,
            cost_usd          REAL,
            currency          TEXT,
            upload_date       TEXT,
            PRIMARY KEY (session_id, resource)
        )
    ''')
    cur.execute('''
        CREATE INDEX IF NOT EXISTS idx_daily_costs_date ON daily_costs(upload_date)
    ''')
    conn.commit()
    cur.close()
    conn.close()
    print('Database initialised.')

# ---------------------------------------------------------------------------
# Classification logic
# ---------------------------------------------------------------------------

AZURE_PROVIDER_TYPE_MAP = {
    'microsoft.compute/virtualmachinescalesets': 'Virtual machine scale set',
    'microsoft.compute/virtualmachines':         'Virtual machine',
    'microsoft.compute/disks':                   'Disk',
    'microsoft.storage/storageaccounts':         'Storage account',
    'microsoft.dbforpostgresql/flexibleservers': 'Azure Database for PostgreSQL flexible server',
    'microsoft.dbformysql/flexibleservers':      'Azure Database for MySQL flexible server',
    'microsoft.network/natgateways':             'NAT gateway',
    'microsoft.network/loadbalancers':           'Load balancer',
    'microsoft.network/privateendpoints':        'Private endpoint',
    'microsoft.network/publicipaddresses':       'Public IP address',
    'microsoft.network/privatednszones':         'Private DNS zone',
}

PLATFORM_STORAGE = {
    'bivasharefolder', 'checkpointsjio', 'bivastoragejio',
    'bivasystemtablesjio', 'bivadbmigration', 'bivajiobilling',
}

PLATFORM_TYPES = {
    'Azure Database for MySQL flexible server',
    'Azure Database for PostgreSQL flexible server',
    'Disk', 'Load balancer', 'NAT gateway', 'Private DNS zone',
    'Private endpoint', 'Public IP address', 'Virtual machine',
}

def normalize_resource_type(rt: str) -> str:
    return AZURE_PROVIDER_TYPE_MAP.get(rt.strip().lower(), rt.strip())

def classify_resource(resource: str, resource_azure_type: str, resource_group: str, db) -> str | None:
    key = resource.strip().lower()
    cur = get_cur(db)
    cur.execute('SELECT type FROM resource_type_map WHERE resource_key = %s', (key,))
    row = cur.fetchone()
    cur.close()
    if row:
        return row['type']

    rt   = normalize_resource_type(resource_azure_type or '')
    name = key

    if rt == 'Virtual machine scale set':
        if 'sparkpool' in name:
            return 'Customer Attributed (Compute)'
        if any(p in name for p in ('bivapool', 'default')):
            return 'Platform'

    if rt == 'Storage account':
        if name in PLATFORM_STORAGE:
            return 'Platform'
        return 'Customer Specific (Storage,Read/write)'

    if rt in PLATFORM_TYPES:
        return 'Platform'

    return None


def parse_rows_from_sheet(ws) -> list[dict]:
    headers = [str(ws.cell(1, c).value or '').strip() for c in range(1, ws.max_column + 1)]
    col = {h.lower(): i + 1 for i, h in enumerate(headers)}

    def gc(row, name, default=None):
        idx = col.get(name.lower())
        return ws.cell(row, idx).value if idx else default

    has_resource_col = 'resource' in col
    has_usage_date   = 'usagedate' in col

    if has_resource_col and not has_usage_date:
        rows = []
        for r in range(2, ws.max_row + 1):
            resource = str(gc(r, 'Resource') or '').strip()
            if not resource:
                continue
            rows.append({
                'resource':          resource,
                'resource_id':       str(gc(r, 'ResourceId') or ''),
                'resource_type':     str(gc(r, 'ResourceType') or ''),
                'resource_group':    str(gc(r, 'ResourceGroupName') or ''),
                'subscription_name': str(gc(r, 'SubscriptionName') or ''),
                'cost_inr':          float(gc(r, 'Cost') or 0),
                'cost_usd':          float(gc(r, 'CostUSD') or 0),
                'currency':          str(gc(r, 'Currency') or 'INR'),
                'file_date':         None,
            })
        return rows
    else:
        agg = {}
        for r in range(2, ws.max_row + 1):
            rid  = str(gc(r, 'ResourceId') or '').strip()
            if not rid:
                continue
            name = rid.split('/')[-1].lower()
            if not name:
                continue
            rt       = str(gc(r, 'ResourceType') or '').strip()
            rg       = str(gc(r, 'ResourceGroupName') or '').strip()
            cost_inr = float(gc(r, 'Cost') or 0)
            cost_usd = float(gc(r, 'CostUSD') or 0)
            currency = str(gc(r, 'Currency') or 'INR')

            raw_date = gc(r, 'UsageDate') or ''
            if hasattr(raw_date, 'date'):
                file_date = raw_date.date().isoformat()
            else:
                file_date = str(raw_date).strip()[:10]

            key = (name, file_date)
            if key not in agg:
                agg[key] = {
                    'resource':          name,
                    'resource_id':       rid,
                    'resource_type':     rt,
                    'resource_group':    rg,
                    'subscription_name': '',
                    'cost_inr':          0.0,
                    'cost_usd':          0.0,
                    'currency':          currency,
                    'file_date':         file_date,
                }
            agg[key]['cost_inr'] += cost_inr
            agg[key]['cost_usd'] += cost_usd

        return list(agg.values())


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@app.route('/')
def index():
    return render_template('index.html', types=TYPES)


@app.route('/upload', methods=['POST'])
def upload():
    file       = request.files.get('file')
    session_id = request.form.get('session_id') or datetime.utcnow().strftime('%Y%m%d%H%M%S%f')

    if not file:
        return jsonify({'error': 'No file provided'}), 400

    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
    except Exception as e:
        return jsonify({'error': f'Cannot read file: {e}'}), 400

    try:
        raw_rows = parse_rows_from_sheet(ws)
    except Exception as e:
        return jsonify({'error': f'Error parsing sheet: {e}'}), 400

    if not raw_rows:
        return jsonify({'error': 'No data rows found in file. Check the file format.'}), 400

    db = get_db()
    classified   = []
    unclassified = []

    for row in raw_rows:
        resource    = row['resource']
        upload_date = row.get('file_date') or date.today().isoformat()
        t = classify_resource(resource, row['resource_type'], row['resource_group'], db)
        entry = {**row, 'upload_date': upload_date}
        entry.pop('file_date', None)
        if t:
            entry['type'] = t
            classified.append(entry)
        else:
            unclassified.append(entry)

    all_dates = sorted(set(r['upload_date'] for r in classified + unclassified))

    cur = get_cur(db)
    existing_by_date = {}
    for d in all_dates:
        cur.execute('SELECT COUNT(*) as cnt FROM daily_costs WHERE upload_date = %s', (d,))
        existing_by_date[d] = cur.fetchone()['cnt']

    if unclassified:
        cur.execute('DELETE FROM pending_classifications WHERE session_id = %s', (session_id,))
        psycopg2.extras.execute_values(cur,
            '''INSERT INTO pending_classifications
               (session_id, resource, resource_id, resource_type, resource_group,
                subscription_name, cost_inr, cost_usd, currency, upload_date)
               VALUES %s''',
            [(session_id, u['resource'], u['resource_id'], u['resource_type'],
              u['resource_group'], u['subscription_name'], u['cost_inr'],
              u['cost_usd'], u['currency'], u['upload_date']) for u in unclassified]
        )
        db.commit()

    cur.close()
    return jsonify({
        'session_id':       session_id,
        'dates':            all_dates,
        'classified_count': len(classified),
        'unclassified':     unclassified,
        'classified':       classified,
        'existing_by_date': existing_by_date,
    })


@app.route('/classify', methods=['POST'])
def classify():
    data       = request.json
    session_id = data.get('session_id')
    selections = data.get('selections', [])

    db  = get_db()
    cur = get_cur(db)

    for sel in selections:
        key = sel['resource'].strip().lower()
        cur.execute(
            '''INSERT INTO resource_type_map (resource_key, type, source)
               VALUES (%s, %s, %s)
               ON CONFLICT (resource_key) DO UPDATE SET type = EXCLUDED.type, source = EXCLUDED.source''',
            (key, sel['type'], 'user')
        )

    cur.execute('DELETE FROM pending_classifications WHERE session_id = %s', (session_id,))
    db.commit()
    cur.close()
    return jsonify({'saved': len(selections)})


@app.route('/commit', methods=['POST'])
def commit():
    data = request.json
    rows = data.get('rows', [])
    if not rows:
        return jsonify({'committed': 0})

    db  = get_db()
    cur = get_cur(db)

    dates_in_batch = set(r['upload_date'] for r in rows)
    for d in dates_in_batch:
        cur.execute('DELETE FROM daily_costs WHERE upload_date = %s', (d,))

    psycopg2.extras.execute_values(cur,
        '''INSERT INTO daily_costs
           (upload_date, resource, resource_id, resource_type, resource_group,
            subscription_name, cost_inr, cost_usd, currency, type)
           VALUES %s''',
        [(r['upload_date'], r['resource'], r['resource_id'], r['resource_type'],
          r['resource_group'], r['subscription_name'], r['cost_inr'],
          r['cost_usd'], r['currency'], r['type']) for r in rows]
    )
    db.commit()
    cur.close()
    return jsonify({'committed': len(rows), 'dates': sorted(dates_in_batch)})


@app.route('/available-dates')
def available_dates():
    db  = get_db()
    cur = get_cur(db)
    cur.execute(
        '''SELECT upload_date, COUNT(*) as rows, SUM(cost_inr) as total
           FROM daily_costs GROUP BY upload_date ORDER BY upload_date DESC'''
    )
    rows = cur.fetchall()
    cur.close()
    return jsonify([dict(r) for r in rows])


@app.route('/report')
def report():
    from_date = request.args.get('from_date')
    to_date   = request.args.get('to_date')
    if not from_date or not to_date:
        days      = int(request.args.get('days', 15))
        to_date   = date.today().isoformat()
        from_date = (date.today() - timedelta(days=days - 1)).isoformat()

    db  = get_db()
    cur = get_cur(db)
    cur.execute(
        '''SELECT upload_date, resource, type, cost_inr
           FROM daily_costs
           WHERE upload_date >= %s AND upload_date <= %s
           ORDER BY upload_date''',
        (from_date, to_date)
    )
    rows  = cur.fetchall()
    cur.close()

    dates = sorted(set(r['upload_date'] for r in rows))

    compute_by_date         = {}
    storage_by_customer_date = {}

    for r in rows:
        d = r['upload_date']
        if r['type'] == 'Customer Attributed (Compute)':
            compute_by_date[d] = compute_by_date.get(d, 0) + r['cost_inr']
        elif r['type'] == 'Customer Specific (Storage,Read/write)':
            k = (r['resource'], d)
            storage_by_customer_date[k] = storage_by_customer_date.get(k, 0) + r['cost_inr']

    daily_chart = []
    for d in dates:
        storage_total = sum(v for (_, dd), v in storage_by_customer_date.items() if dd == d)
        daily_chart.append({
            'date':    d,
            'storage': round(storage_total, 2),
            'compute': round(compute_by_date.get(d, 0), 2),
        })

    customers = sorted(set(k[0] for k in storage_by_customer_date.keys()))
    table = []
    for customer in customers:
        total_storage = 0
        total_compute_apportioned = 0
        for d in dates:
            s             = storage_by_customer_date.get((customer, d), 0)
            total_storage += s
            total_compute = compute_by_date.get(d, 0)
            total_storage_day = sum(
                v for (_, dd), v in storage_by_customer_date.items() if dd == d
            )
            if total_storage_day > 0 and total_compute > 0:
                total_compute_apportioned += total_compute * (s / total_storage_day)
        table.append({
            'customer':     customer,
            'storage_cost': round(total_storage, 2),
            'compute_cost': round(total_compute_apportioned, 2),
            'total_cost':   round(total_storage + total_compute_apportioned, 2),
        })

    table.sort(key=lambda x: -x['total_cost'])

    return jsonify({
        'dates':       dates,
        'from_date':   from_date,
        'to_date':     to_date,
        'daily_chart': daily_chart,
        'table':       table,
        'totals': {
            'storage_cost': round(sum(r['storage_cost'] for r in table), 2),
            'compute_cost': round(sum(r['compute_cost'] for r in table), 2),
            'total_cost':   round(sum(r['total_cost'] for r in table), 2),
        }
    })


@app.route('/history')
def history():
    db  = get_db()
    cur = get_cur(db)
    cur.execute(
        '''SELECT upload_date, COUNT(*) as rows, SUM(cost_inr) as total
           FROM daily_costs GROUP BY upload_date ORDER BY upload_date DESC'''
    )
    rows = cur.fetchall()
    cur.close()
    return jsonify([dict(r) for r in rows])


if DATABASE_URL:
    try:
        init_db()
    except Exception as e:
        print(f'Warning: init_db failed: {e}')

if __name__ == '__main__':
    app.run(debug=True, port=5000)
