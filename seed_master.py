"""
Seed the resource_type_map table from the master Excel sheet.
Run once: python seed_master.py
"""
import os
import openpyxl
import psycopg2
from dotenv import load_dotenv

load_dotenv()

MASTER       = os.path.join(os.path.dirname(__file__), 'data', 'master.xlsx')
DATABASE_URL = os.environ.get('DATABASE_URL')

def seed(xlsx_path=MASTER):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    headers = [str(ws.cell(1, c).value or '').strip() for c in range(1, ws.max_column + 1)]
    col = {h.lower(): i + 1 for i, h in enumerate(headers)}

    conn = psycopg2.connect(DATABASE_URL)
    cur  = conn.cursor()
    count = 0
    for r in range(2, ws.max_row + 1):
        resource = str(ws.cell(r, col.get('resource', 1)).value or '').strip()
        t        = str(ws.cell(r, col.get('type', ws.max_column)).value or '').strip()
        if resource and t:
            cur.execute(
                '''INSERT INTO resource_type_map (resource_key, type, source)
                   VALUES (%s, %s, %s)
                   ON CONFLICT (resource_key) DO UPDATE SET type = EXCLUDED.type, source = EXCLUDED.source''',
                (resource.lower(), t, 'master')
            )
            count += 1
    conn.commit()
    cur.close()
    conn.close()
    print(f'Seeded {count} resource→type mappings from master.')

if __name__ == '__main__':
    seed()
